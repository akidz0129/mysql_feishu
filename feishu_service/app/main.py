
import re
import os
import json
import httpx
import logging
import asyncio
import pandas as pd
from io import BytesIO
import lark_oapi as lark
from lark_oapi.api.im.v1 import *
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
# --- 模块级别日志设置（优先执行） ---
# 获取当前模块的 logger 实例
logger = logging.getLogger(__name__) 
# 创建一个 StreamHandler，将日志输出到控制台
handler = logging.StreamHandler()
# 设置日志格式
formatter = logging.Formatter('%(asctime)s %(name)-8s %(levelname)-8s %(message)s [%(filename)s:%(lineno)d]')
handler.setFormatter(formatter)
# 为 logger 添加处理器
logger.addHandler(handler)
# 设置日志级别（例如 INFO，这样 INFO, WARNING, ERROR, CRITICAL 都会显示）
logger.setLevel(logging.INFO)

APP_ID=os.environ.get("FEISHU_APP_ID")
APP_SECRET=os.environ.get("FEISHU_APP_SECRET")
FOLDER_TOKEN=os.environ.get("FEISHU_FOLDER_TOKEN")

async def tenant_access_token(app_id, app_secret):
    url = 'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal'
    headers = {'Content-Type':'application/json; charset=utf-8'}
    json_body = {'app_id': app_id, 'app_secret': app_secret}

    async with httpx.AsyncClient() as client:
        resp = await client.post(url, headers=headers, json=json_body)
        res = resp.json()
        print(res)
        return res['tenant_access_token']

MAPPING={
'ID':                                                   'ID',
'Purchase_Datetime':                                    '订购日期',  
'Order_ID':                                             '订单号',  
'Reference_ID':                                         '参考号',                                        
'Region':                                               '大区',
'Ship_Country_Name':                                    '国家名',       
'Ship_Country':                                         '国家', 
'Currency':                                             '原币种',
'Exchange_Rate':                                        '汇率',
'Fulfillment_Channel':                                  '渠道',                      
'Store':                                                '店铺',                 
'Tags':                                                 '父ASIN/spu',         
'Product_ID':                                           '子ASIN/sku',         
'MSKU':                                                 'MSKU',             
'MSKU_True':                                            'MSKU_True',
'Item_ID':                                              '商品ID'                                        ,
'GTIN':                                                 'GTIN',
'Product_Chinese_Name':                                 '中文品名',
'Product_Name':                                         '品名',           
'Variation_Name':                                       '规格',
'Quantity':                                             '销量',    
'Item_Price':                                           '商品金额', 
'Buyer_Paid_Amount':                                    '买家实付金额'                      ,
'Sale_Amount':                                          '实际销售额'                    ,
'Repayment_Amount':                                     '手动计算回款额'          ,
'Profit':                                               '最终利润',             
'Seller_Voucher':                                       '卖家优惠券金额', 
'Platform_Voucher':                                     '平台优惠券', 
'Coin_Offset':                                          '币抵扣',
'Buyer_Paid_Shipping_Fee':                              '买家支付运费', 
'Seller_Paid_Shipping_Fee':                             '卖家支付运费'                            ,
'Actual_Shipping_Fee':                                  '实际物流运费', 
'Shipping_Subsidy':                                     '运费返还', 
'Platform_Discount':                                    '平台回扣',
'Initial_Buyer_TXN_Fee':                                '买家初始交易费',
'Buyer_Service_Fee':                                    '买家服务费',
'Insurance_Premium':                                    '保险费',
'Shipping_Fee_SST_Amount':                              '运费SST',
'Item_Tax':                                             '税费',
'Affiliate_Marketing':                                  '联盟营销方案佣金', 
'Commission_Fee':                                       '佣金',
'Service_Fee':                                          '服务费', 
'Transaction_Fee':                                      '交易费',
'Delivery_Seller_Protection_Fee_Premium_Amount':        '配送卖家保护费高级金额',
'Seller_Coin_Cash_Back':                                '卖家币返还',
'Seller_Order_Processing_Fee':                          '订单处理费',
'Vat':                                                  '增值税',
'Return_Quantity':                                      '退货数量',
'Return_Price':                                         '退款金额',
'Return_Credit':                                        '退款入账',
'Return_Debit':                                         '退款支出',          
'Product_Cost':                                         '产品成本',  
'First_Leg_Shipping_Cost':                              '头程',
'Outbound_Cost':                                        '出库费',                              
'Shipping_Cost':                                        '运费',      
'Advertising_Spend':                                    '广告花费',
'Lost_Claim':                                           '丢件索赔',
'Other_Debit':                                          '其它支出',
'Influencer_Commission_Fee':                            '达人佣金'                        ,
'Influencer_Partner_Commission_Fee':                    '达人合作伙伴佣金'        ,
'SFP_Service_Fee':                                      'SFP服务费'                                 ,
'Bonus_Cashback_Service_Fee':                           'Bonus Cashback Service Fee'     ,
'Profit_Adjustment':                                    '补毛保'                                  ,
'Invoice_Fee':                                          '开票费用'                                      ,
'Amortization':                                         '摊销'                                        ,
'Ship_Promotion_Discount':                              '促销费-运费折扣'                   ,
'Item_Promotion_Discount':                              '促销费-商品折扣'                   ,
'FBA_Fee':                                              'FBA费'                                             ,
'Promotion_Discount':                                   '促销折扣'                               ,
'Inventory_Reimbursement':                              '库存赔偿'                          ,
'Other_Credit':                                         '其他收入',
'Fee_Refund_Amount':                                    '费用退款额'                              ,
'Promotion_Cost':                                       '推广费'                                     ,
'Storage_Fee':                                          '仓储费'                                        ,
'Inbound_Cost':                                         '入库配置费'                                   ,
'Adjustment_Cost':                                      '调整费用'                                  ,
'Removal_Cost':                                         '移除费'                                       ,
'Transaction_Service_Fee':                              '交易服务费'                        ,
'JD_Paid_Global_Sale_Tax':                              '京东代付全球售消费税'              ,
'JD_Collected_Global_Sale_Tax':                         '京东代收全球售消费税'         ,
'JD_Points':                                            '京豆'                                           ,
'Collected_Shipping_Fee':                               '代收配送费'                         ,
'Price_Protection_Deduction':                           '价保扣款'                       ,
'Price_Protection_Rebate':                              '价保返佣'                          ,
'Seller_Shipping_Return':                               '卖家返还运费'                           ,
'After_Sale_Seller_Compensation':                       '售后卖家赔付费'                 ,
'Product_Insurance_Fee':                                '商品保险服务费'                          ,
'Direct_Compensation_Fee':                              '商家直赔费'                            ,
'Small_Payment_Refund':                                 '小额打款费用退还'                         ,
'Platform_Coupon_Subsidy':                              '平台券价保补贴'                        ,
'Platform_Coupon_Subsidy_Commission':                   '平台券价保补贴佣金'         ,
'Ad_Cooperation_Deduction_Commission':                  '广告联合活动降扣佣金'      ,
'Comprehensive_Penalty':                                '综合违约金'                              ,
'Shipping_Insurance_Fee':                               '运费保险服务费'                         ,
'SOP_Double_Compensation_JDPoints':                     'SOP双倍赔款（京豆赔付）'      ,
'SOP_Double_Compensation_Commission':                   'SOP双倍赔返还佣金'          ,
'JD_Promotion_Tech_Service_Fee':                        '代收白条网络推广技术服务费'                ,
'NonPlatform_National_Subsidy_Transaction_Fee':         '非平台核销国补交易服务费'   ,
'NonPlatform_National_Subsidy_Commission':              '非平台核销国补佣金'             ,
'NonPlatform_National_Subsidy_Ad_Commission':           '非平台核销国补广告佣金'       ,
}
FIELD_MAPPING = {v: k for k, v in MAPPING.items()}
fastapi_order_url = "http://fastapi-app:8000/orders"
async_client = httpx.AsyncClient(timeout=10.0)

def statistic_table(stream):
    raw_df = pd.read_excel(stream, sheet_name='Sheet', engine='openpyxl') # 默认 Sheet1, header=0
    rename_map={
        '中文品名':'品名',
        'MSKU':'69码',
        '商品金额':'商品金额',
        '买家实付金额':'买家实付',
        '手动计算回款额':'回款金额',
        '最终利润':'毛利润',
        '联盟营销方案佣金':'联盟营销',
    }
    ALLOWED_DB_COLUMNS=[
        '店铺','国家','父ASIN/spu','子ASIN/sku','品名','69码','访客','销量','订单件数',
        '商品金额','实际销售额','买家实付','回款金额','产品成本','毛利润','销售利润率','回款利润率',
        '佣金','交易费','服务费','头程','出库费','运费','联盟营销','广告花费','退货数量','退货率','退款支出','退款金额'
    ]

    df=raw_df[['渠道','店铺','国家名','父ASIN/spu','子ASIN/sku','中文品名','MSKU',
        '销量','商品金额','买家实付金额','手动计算回款额','产品成本','最终利润','佣金','交易费','服务费','头程','出库费',
        '运费','联盟营销方案佣金','广告花费','退货数量','退款金额','实际销售额','退款支出'
    ]].rename(columns={'国家名':'国家','销量':'订单件数'})
    meta_cols = ['中文品名','MSKU']
    sum_cols = ['订单件数','商品金额','实际销售额','买家实付金额','手动计算回款额','产品成本','最终利润','佣金','交易费','服务费','头程','出库费','运费','联盟营销方案佣金','广告花费','退货数量','退款金额','退款支出']

    for s in sum_cols:
        df.loc[:, s] = pd.to_numeric(df[s], errors='coerce').fillna(0)
    # ------------shopee--------------
    grouped = df.loc[df['渠道'].astype(str).isin(['shopee'])].groupby(['店铺','国家','父ASIN/spu','子ASIN/sku'], as_index=False).agg(
        {**{col:'first' for col in meta_cols}, **{col:'sum' for col in sum_cols}}
    )
    # grouped =df.groupby(['父ASIN/spu','子ASIN/sku'],as_index=False)[['销量','手动计算回款额','最终利润']].sum()
    row_counts = df.loc[df['渠道'].astype(str).isin(['shopee'])].groupby(['店铺','国家','父ASIN/spu','子ASIN/sku'], as_index=False).size().rename(columns={'size':'销量'})

    totals=grouped.groupby(['店铺','国家','父ASIN/spu'],as_index=False).agg(
        {**{col:'first' for col in meta_cols}, **{col:'sum' for col in sum_cols}}
    )
    row_countss = df.groupby(['店铺','国家','父ASIN/spu'], as_index=False).size().rename(columns={'size':'销量'})

    totals['子ASIN/sku'] = '-'

    grouped = grouped.merge(row_counts, on=['店铺','国家','父ASIN/spu','子ASIN/sku'])
    totals = totals.merge(row_countss, on=['店铺','国家','父ASIN/spu'])
    final_df = pd.concat([grouped, totals], ignore_index=True)

    # -----------------lazada_tiktok----------------
    grouped_lt = df.loc[df['渠道'].astype(str).isin(['lazada','tiktok'])].groupby(['店铺','国家','MSKU'], as_index=False).agg(
        {**{col:'first' for col in meta_cols}, **{col:'sum' for col in sum_cols}}
    )
    row_counts_lt = df.loc[df['渠道'].astype(str).isin(['lazada','tiktok'])].groupby(['店铺','国家','MSKU'], as_index=False).size().rename(columns={'size':'销量'})

    grouped_lt['子ASIN/sku'] = '-'
    grouped_lt['父ASIN/spu'] = '-'

    grouped_lt = grouped_lt.merge(row_counts_lt, on=['店铺','国家','MSKU'])

    final_df = pd.concat([final_df, grouped_lt], ignore_index=True)

    # -----------------amazon----------------
    grouped_a = df.loc[df['渠道'].astype(str).isin(['amazon'])].groupby(['店铺','国家','父ASIN/spu'], as_index=False).agg(
        {**{col:'first' for col in meta_cols}, **{col:'sum' for col in sum_cols}}
    )
    # grouped =df.groupby(['父ASIN/spu','子ASIN/sku'],as_index=False)[['销量','手动计算回款额','最终利润']].sum()
    row_counts_a = df.loc[df['渠道'].astype(str).isin(['amazon'])].groupby(['店铺','国家','父ASIN/spu'], as_index=False).size().rename(columns={'size':'销量'})

    grouped_a['子ASIN/sku'] = '-'
    grouped_a = grouped_a.merge(row_counts_a, on=['店铺','国家','父ASIN/spu'])

    final_df = pd.concat([final_df, grouped_a], ignore_index=True)
    # -----------------walmart----------------

    grouped_w = df.loc[df['渠道'].astype(str).isin(['walmart'])].groupby(['店铺','国家','父ASIN/spu','子ASIN/sku'], as_index=False).agg(
        {**{col:'first' for col in meta_cols}, **{col:'sum' for col in sum_cols}}
    )
    row_counts_w = df.loc[df['渠道'].astype(str).isin(['walmart'])].groupby(['店铺','国家','父ASIN/spu','子ASIN/sku'], as_index=False).size().rename(columns={'size':'销量'})

    grouped_w = grouped_w.merge(row_counts_w, on=['店铺','国家','父ASIN/spu','子ASIN/sku'])
    final_df = pd.concat([final_df, grouped_w], ignore_index=True)

    # -----------------jd----------------

    grouped_j = df.loc[df['渠道'].astype(str).isin(['jd_self','jd_pop'])].groupby(['店铺','国家','子ASIN/sku'], as_index=False).agg(
        {**{col:'first' for col in meta_cols}, **{col:'sum' for col in sum_cols}}
    )
    row_counts_j = df.loc[df['渠道'].astype(str).isin(['jd_self','jd_pop'])].groupby(['店铺','国家','子ASIN/sku'], as_index=False).size().rename(columns={'size':'销量'})

    grouped_j = grouped_j.merge(row_counts_j, on=['店铺','国家','子ASIN/sku'])
    final_df = pd.concat([final_df, grouped_j], ignore_index=True)

    # --------------------------------
    final_df['sort_order'] = final_df['子ASIN/sku'].apply(lambda x: 0 if x=='-' else 1)
    final_df = final_df.sort_values(['父ASIN/spu', 'sort_order', '子ASIN/sku'], ignore_index=True)
    final_df = final_df.drop(columns='sort_order').rename(columns=rename_map)
    final_df['毛利润'] = final_df['毛利润'].astype(float)
    final_df['实际销售额'] = final_df['实际销售额'].astype(float)
    import numpy as np
    with np.errstate(divide='ignore', invalid='ignore'):
        result = np.divide(final_df['毛利润'], final_df['实际销售额'])

    final_df['销售利润率'] = np.where(
        final_df['实际销售额'] == 0, 
        0,                             
        result                         
    )

    with np.errstate(divide='ignore', invalid='ignore'):
        result = np.divide(final_df['毛利润'], final_df['回款金额'])

    final_df['回款利润率']= np.where(
        final_df['回款金额'] == 0,
        0,  
        result
    )

    final_df['退货数量'] = pd.to_numeric(final_df['退货数量'], errors='coerce')
    final_df['订单件数'] = pd.to_numeric(final_df['订单件数'], errors='coerce')

    with np.errstate(divide='ignore', invalid='ignore'):
        # 在这个块内部执行除法，NumPy 将不会对除以零抛出异常
        result = np.divide(final_df['退货数量'], final_df['订单件数'])
    final_df['退货率'] = np.where(
        final_df['订单件数'] == 0, 
        0,                             
        result                         
    )



    final_df=final_df.reindex(columns=ALLOWED_DB_COLUMNS)
    # ----------------- 内存中的写出和格式化 -----------------

    stream.seek(0)
    workbook = load_workbook(stream)
    # 3. 创建新的 Sheet 页并写入数据
    new_sheet_name = '统计表'
    # 检查是否已存在，如果存在则删除/替换
    if new_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[new_sheet_name])
    worksheet = workbook.create_sheet(new_sheet_name)
    output_stream = BytesIO() 
    
    for r_idx, r in enumerate(dataframe_to_rows(final_df, header=True, index=False)):
        worksheet.append(r)
        

        # 定义表头格式
    header_fill = PatternFill(start_color="F2C150", end_color="F2C150", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, col_name in enumerate(final_df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    # 2. 设置百分比格式 
    percent_cols = [16, 17, 27] 
    
    for col in percent_cols:
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = '0.00%' 

    workbook.save(output_stream)
    # 3. 结果返回
    output_stream.seek(0) # 将指针重置到流的开头
    
    # 返回 BytesIO 对象本身，或者返回其内容 (bytes)，取决于您的调用方需求。
    return output_stream

def json_to_excel_stream(json_data):
    wb = Workbook()
    ws = wb.active

    # 如果 json_data 是列表（通常是多行数据）
    if isinstance(json_data, list):

        # 写表头        
        if isinstance(json_data[0], dict):
            original_keys = list(json_data[0].keys())
            headers = [MAPPING.get(k, k) for k in original_keys]
            ws.append(headers)
            # ws.append(list(json_data[0].keys()))

        # 写数据行
        for row in json_data:
            if isinstance(row, dict):
                ws.append(list(row.values()))
            else:
                ws.append(row)

    # 单行数据
    elif isinstance(json_data, dict):
        ws.append(list(json_data.keys()))
        ws.append(list(json_data.values()))

    # 输出到内存
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    stream = statistic_table(stream)
    print(len(stream.getvalue()))
    return stream, len(stream.getvalue())

# 分片上传文件至飞书
async def upload_file_feishu(access_token,file_stream, file_size, parent_node, file_name):
    # Step 1: prepare
    prepare_url = "https://open.feishu.cn/open-apis/drive/v1/files/upload_prepare"
    prepare_body = {
        "file_name": file_name,
        "parent_node": parent_node,
        "parent_type": "explorer",
        "size": file_size
    }
    headers = {"Authorization": f"Bearer {access_token}"}

    async with httpx.AsyncClient() as client:
        prep = await client.post(prepare_url, json=prepare_body, headers=headers)

        prep_data = prep.json()["data"]
        logger.info(prep_data)

    upload_id = prep_data["upload_id"]
    block_size = prep_data["block_size"]
    block_num = prep_data["block_num"]

    # Step 2. upload parts
    upload_url = "https://open.feishu.cn/open-apis/drive/v1/files/upload_part"
    file_stream.seek(0)
    async with httpx.AsyncClient() as client:
        for seq in range(block_num):
            chunk = file_stream.read(block_size)
            logger.info(f"upload seq {seq}, size={len(chunk)}")

            files = {
                "file": ("blob", chunk)
            }
            data = {
                "upload_id": upload_id,
                "seq": seq,
                "size": len(chunk)
            }

            await client.post(upload_url, data=data, files=files, headers=headers)
            logger.info(f"prep status = {prep.status_code}")
            logger.info(f"prep text = {prep.text}")


    # Step 3: finish
    finish_url = "https://open.feishu.cn/open-apis/drive/v1/files/upload_finish"
    async with httpx.AsyncClient() as client:
        fin = await client.post(finish_url, json={"upload_id": upload_id,"block_num":block_num}, headers=headers)

    return fin.json()['data']['file_token']

async def daily_report(start_date: str, end_date: str, fulfillment_channel: str, msku: str, columns: list[str]) -> str:
    """
    向 FastAPI 服务发送日报查询请求。
    """
    payload = {
        "start_date": start_date,
        "end_date": end_date,
        "fulfillment_channel":fulfillment_channel,
        "msku":msku,
        "columns": columns,
    }
    logger.info(payload)

    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(f"{fastapi_order_url}/report", json=payload, timeout=30.0)
            response.raise_for_status() # 检查 HTTP 状态码

            fastapi_response_data = response.json()
            # 假设 FastAPI 返回的数据结构类似 {"status": "success", "data": [...]}
            if fastapi_response_data.get("status") == "success":
                report_data = fastapi_response_data.get("data", [])
                if report_data:
                    file_stream, file_size=json_to_excel_stream(report_data)
                    FEISHU_APP_TOKEN = await tenant_access_token(APP_ID,APP_SECRET)
                    logger.info(f'{start_date}_{end_date}.xlsx')


                    file_token = await upload_file_feishu(FEISHU_APP_TOKEN,file_stream, file_size, FOLDER_TOKEN, f'{fulfillment_channel}{start_date}_{end_date}.xlsx')
                    logger.info("开始 prepare 上传4…")
                    # 格式化查询结果为可读的字符串
                    # for row in report_data:
                    #     formatted_report += f"- {row}\n" # 假设 row 是一个字典或列表
                    formatted_report=f'https://www.feishu.cn/file/{file_token}'
                    return formatted_report
                else:
                    return f"未找到 {start_date}_{end_date} 的日报数据。"
            else:
                error_msg = fastapi_response_data.get("message", "未知错误")
                return f"日报查询失败: {error_msg}"
    except httpx.HTTPStatusError as e:
        return f"查询日报失败，服务器返回错误：{e.response.status_code}。请稍后再试。"
    except httpx.RequestError as e:
        return "查询日报失败，无法连接到查询服务。请检查服务状态和FASTAPI_ORDER_URL。"
    except Exception as e:
        return "查询日报时发生未知错误，请联系管理员。"

async def daily_item(start_date: str, end_date: str, fulfillment_channel: str, msku: str, columns: list[str]) -> str:
    """
    向 FastAPI 服务发送商品查询请求。
    """
    payload = {
        "start_date": start_date,
        "end_date": end_date,
        "fulfillment_channel":fulfillment_channel,
        "msku":msku,
        "columns": columns,
    }
    logger.info(payload)

    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(f"{fastapi_order_url}/item_info", json=payload, timeout=30.0)
            response.raise_for_status() # 检查 HTTP 状态码

            fastapi_response_data = response.json()
            # 假设 FastAPI 返回的数据结构类似 {"status": "success", "data": [...]}
            if fastapi_response_data.get("status") == "success":
                report_data = fastapi_response_data.get("data", [])
                if report_data:
                    formatted_report = f"{start_date}_{end_date}_{fulfillment_channel}_{msku}:\n"
                    for k, v in report_data.items():
                        formatted_report += f"- {MAPPING.get(k)}: {v}\n"
                    return formatted_report
                else:
                    return f"未找到 {start_date}_{end_date} 的数据。"
            else:
                error_msg = fastapi_response_data.get("message", "未知错误")
                return f"日报查询失败: {error_msg}"
    except httpx.HTTPStatusError as e:
        return f"查询日报失败，服务器返回错误：{e.response.status_code}。请稍后再试。"
    except httpx.RequestError as e:
        return "查询日报失败，无法连接到查询服务。请检查服务状态和FASTAPI_ORDER_URL。"
    except Exception as e:
        return "查询日报时发生未知错误，请联系管理员。"

# ------------------- 异步发送飞书消息 -------------------
async def send_feishu_text(open_id: str, text: str):
    url = "https://open.feishu.cn/open-apis/im/v1/messages"
    FEISHU_APP_TOKEN = await tenant_access_token(APP_ID,APP_SECRET)
    headers = {
        "Authorization": f"Bearer {FEISHU_APP_TOKEN}",
        "Content-Type": "application/json; charset=utf-8"
    }
    params = {"receive_id_type":"chat_id"}
    body={
        "receive_id": open_id,
        "content": json.dumps({"text": text}),
        "msg_type": "text"
    }
    resp = await async_client.post(url, headers=headers, params=params,json=body)
    data = resp.json()
    if data.get("code") != 0:
        print("发送失败:", data)
    return data
# ------------------- 异步发送飞书富文本消息（支持加粗） -------------------
async def send_feishu_post(open_id: str, title: str, content_blocks: list):
    """
    content_blocks 示例：
    [
        [
        ]
    ]
    """
    url = "https://open.feishu.cn/open-apis/im/v1/messages"
    FEISHU_APP_TOKEN = await tenant_access_token(APP_ID, APP_SECRET)

    headers = {
        "Authorization": f"Bearer {FEISHU_APP_TOKEN}",
        "Content-Type": "application/json; charset=utf-8"
    }

    params = {"receive_id_type": "chat_id"}

    body = {
        "receive_id": open_id,
        "receive_id_type": "chat_id",
        "msg_type": "post",
        "content": json.dumps({
            "zh_cn": {
                "title": title,
                "content": content_blocks
            }
        })
    }

    resp = await async_client.post(url, headers=headers, params=params, json=body)
    data = resp.json()

    if data.get("code") != 0:
        print("发送失败:", data)
    return data

# ------------------- 异步调用 FastAPI -------------------
async def query_daily_report(report_date: str, columns: list[str]) -> str:
    payload = {"report_date": report_date, "columns": columns}
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(f"{fastapi_order_url}/report", json=payload, timeout=30.0)
            resp.raise_for_status()
            data = resp.json()
            if data.get("status") == "success":
                report_data = data.get("data", [])
                if report_data:
                    formatted = f"{report_date} 日报结果:\n"
                    for row in report_data[:5]:  # 仅返回前5条示例
                        formatted += f"- {row}\n"
                    return formatted
                else:
                    return f"未找到 {report_date} 的日报数据。"
            else:
                return f"日报查询失败: {data.get('message','未知错误')}"
    except Exception as e:
        return f"查询日报失败: {e}"

# ------------------- 异步消息处理 -------------------
async def handle_message(data: lark.im.v1.P2ImMessageReceiveV1):
    try:
        content = json.loads(data.event.message.content)
        user_input = content.get("text", "")
    except:
        user_input = "解析消息失败"

    reply_text = "听不懂，请输入帮助查询"

    if user_input=="帮助":
        reply_text= """我可以为您提供以下服务："""


        lines = [
        [{"tag": "text", "text": "1. 查询特定日期报告: ", "style": ["bold"]}],
        [{"tag": "text", "text": "#报告#YYYY-MM-DD#平台:xx#"}],
        [{"tag": "text", "text": "2. 查询特定日期 + 特定项报告: ", "style": ["bold"]}],
        [{"tag": "text", "text": "#报告#YYYY-MM-DD#平台:xx#项名#项名#"}],
        [{"tag": "text", "text": "3. 查询特定日期 + 特定商品（msku 为 69 码，目前仅支持 69 码）报告: ", "style": ["bold"]}],
        [{"tag": "text", "text": "#报告#YYYY-MM-DD#平台:xx#msku:xxx#"}],
        [{"tag": "text", "text": "4. 查询特定日期 + 特定项 + 特定商品（msku 为 69 码，目前仅支持 69 码）报告: ", "style": ["bold"]}],
        [{"tag": "text", "text": "#报告#YYYY-MM-DD#平台:xx#msku:xxx#项名#项名#"}],

        [{"tag": "hr"}],
        [{"tag": "text", "text": "5. 查询特定日期 + 特定项 + 特定商品（msku 为 69 码，目前仅支持 69 码）概览: ", "style": ["bold"]}],
        [{"tag": "text", "text": "#商品#YYYY-MM-DD#平台:xx#msku:xxx#项名#项名#"}],
        [{"tag": "hr"}],
        [{"tag": "text", "text": "6. 查询日期范围报告(1-4均可): ", "style": ["bold"]}],
        [{"tag": "text", "text": "#报告#YYYY-MM-DD#YYYY-MM-DD#平台:xx#"}],
        [{"tag": "text", "text": "7. 查询日期范围 + 特定项 + 特定商品（msku 为 69 码，目前仅支持 69 码）概览: ", "style": ["bold"]}],
        [{"tag": "text", "text": "#商品#YYYY-MM-DD#YYYY-MM-DD#平台:xx#msku:xxx#项名#项名#"}],
        [{"tag": "hr"}],
        [{"tag": "text", "text": "输入：可查询项\n即可查看所有支持的字段。"}],
        [{"tag": "text", "text": "如需进一步帮助，请随时告诉我！"}]
    ]

# 调用发送函数示例
        await send_feishu_post(data.event.message.chat_id, '帮助菜单', lines)
        return
    if user_input=='可查询项':
        reply_text='可查询项有:\n'
        for k, v in MAPPING.items():
            reply_text += f"- {v}\n"

    # 匹配日报请求
    # detail_date_report_match=re.match(r"^#\s*日报#(\d{4}-\d{1,2}-\d{1,2})#([^#]+)#(.*)$", user_input.strip(), re.IGNORECASE)
    report_match=re.match(r"^#\s*报告#(\d{4}-\d{1,2}-\d{1,2})#(\d{4}-\d{1,2}-\d{1,2})#(.*)$", user_input.strip(), re.IGNORECASE)
    daily_report_match=re.match(r"^#\s*报告#(\d{4}-\d{1,2}-\d{1,2})#(.*)$", user_input.strip(), re.IGNORECASE)
    if report_match or daily_report_match:
        if report_match:
            start_date_str = report_match.group(1)
            end_date_str = report_match.group(2)
            remaining_content = report_match.group(3)
            
        elif daily_report_match:
            start_date_str = end_date_str = daily_report_match.group(1)
            remaining_content = daily_report_match.group(2)
        parts = remaining_content.split('#')
        
        msku = None  # ⭐ 默认值：无商品
        fulfillment_channel = None
        requested_db_fields = []

        for part in parts:
            if not part:
                continue

            # 匹配MSKU（格式：MSKU:123456）
            if part.lower().startswith("msku:"):
                msku = part.split(":", 1)[1].strip() or None  # ⭐ 空值也变成 None
                continue
            if part.lower().startswith("平台:"):
                fulfillment_channel = part.split(":", 1)[1].strip() or None  # ⭐ 空值也变成 None
                continue
            logger.info(part)

            # ② 正常字段映射
            mapped_field = FIELD_MAPPING.get(part)
            if mapped_field:
                requested_db_fields.append(mapped_field)
                logger.info(mapped_field)
        # 调 FastAPI

        response_message = await daily_report(
            start_date = start_date_str,
            end_date = end_date_str,
            fulfillment_channel = fulfillment_channel,
            msku = msku,
            columns = requested_db_fields,
        )

        reply_text = response_message
    item_match=re.match(r"^#\s*商品#(\d{4}-\d{1,2}-\d{1,2})#(\d{4}-\d{1,2}-\d{1,2})#(.*)$", user_input.strip(), re.IGNORECASE)
    daily_item_match=re.match(r"^#\s*商品#(\d{4}-\d{1,2}-\d{1,2})#(.*)$", user_input.strip(), re.IGNORECASE)
    # detail_date_item_match=re.match(r"^#\s*商品#(\d{4}-\d{1,2}-\d{1,2})#([^#]+)#([^#]+)#(.*)$", user_input.strip(), re.IGNORECASE)
    if item_match or daily_item_match:
        if item_match:
            start_date_str = item_match.group(1)
            end_date_str = item_match.group(2)
            remaining_content = item_match.group(3)
        elif daily_item_match:
            start_date_str = end_date_str = daily_item_match.group(1)
            remaining_content = daily_item_match.group(2)

        parts = remaining_content.split('#')
        msku = None  # ⭐ 默认值：无商品
        fulfillment_channel = None
        requested_db_fields = []

        for part in parts:
            if not part:
                continue

            # 匹配MSKU（格式：MSKU:123456）
            if part.lower().startswith("msku:"):
                msku = part.split(":", 1)[1].strip() or None  # ⭐ 空值也变成 None
                continue
            if part.lower().startswith("平台:"):
                fulfillment_channel = part.split(":", 1)[1].strip() or None  # ⭐ 空值也变成 None
                continue
            logger.info(part)

            # ② 正常字段映射
            mapped_field = FIELD_MAPPING.get(part)
            if mapped_field:
                requested_db_fields.append(mapped_field)
                logger.info(mapped_field)
        
                # 调用 FastAPI 订单服务
        response_message = await daily_item(
            start_date = start_date_str,
            end_date = end_date_str,
            fulfillment_channel = fulfillment_channel,
            msku = msku,
            columns = requested_db_fields,
        )
        reply_text = response_message
    # 其他匹配逻辑可以按需添加

    # 发送异步回复
    await send_feishu_text(data.event.message.chat_id, reply_text)
    # await send_feishu_text(data.event.sender.sender_id.open_id, reply_text)

# ------------------- 同步回调调度异步任务 -------------------
def do_p2_im_message_receive_v1(data):
    asyncio.create_task(handle_message(data))

# ------------------- 事件处理 -------------------
def do_message_event(data: lark.CustomizedEvent):
    print(f"[customized event] {lark.JSON.marshal(data, indent=4)}")

event_handler = lark.EventDispatcherHandler.builder(APP_ID, APP_SECRET) \
    .register_p2_im_message_receive_v1(do_p2_im_message_receive_v1) \
    .register_p1_customized_event("自定义事件key", do_message_event) \
    .build()

# ------------------- 启动机器人 -------------------
def main():
    cli = lark.ws.Client(APP_ID, APP_SECRET, event_handler=event_handler, log_level=lark.LogLevel.DEBUG)
    cli.start()

if __name__ == "__main__":
    main()












# import lark_oapi as lark
# import json
# from lark_oapi.api.im.v1 import *
# import re
# import os
# import datetime


# ## P2ImMessageReceiveV1 为接收消息 v2.0；CustomizedEvent 内的 message 为接收消息 v1.0。
# APP_ID=os.environ.get("FEISHU_APP_ID")
# APP_SECRET=os.environ.get("FEISHU_APP_SECRET")
# ## P2ImMessageReceiveV1：
# ## P2 对应 v2.0 版本，
# ## ImMessageReceiveV1 对应接收消息事件类型 im.message.receive_v1。
# def do_p2_im_message_receive_v1(data: lark.im.v1.P2ImMessageReceiveV1) -> None:
#     user_input = ""
#     if data.event.message.message_type == "text":
#         user_input = json.loads(data.event.message.content)["text"]
#     else:
#         user_input = "解析消息失败，请发送文本消息\nparse message failed, please send text message"
#     info = ""
    

#     normalized_input = user_input.strip().upper()
#     if normalized_input in ["帮助", "HELP"]:
#             info = """我可以为您提供以下服务:
# - 特定日期日报: 请尝试格式如 “#日报#YYYY-MM-DD#”
# - 特定日期特定产品销量: 请尝试格式如 “#销量#YYYY-MM-DD#asin#”
# - 特定日期特定产品范围销量: 请尝试格式如 “#销量#YYYY-MM-DD(起始时间)#YYYY-MM-DD(结束时间)#asin#”
# - 特定日期销量排名: 请尝试格式如 “#排名#YYYY-MM-DD#”
# - 特定日期范围销量排名: 请尝试格式如 “#排名#YYYY-MM-DD(起始时间)#YYYY-MM-DD(结束时间)#”
# 如果您有其他需求，也可以尝试告诉我。
# """
#     # 匹配意图
#     date_error="日期格式不正确。请使用 YYYY-MM-DD 格式，例如：#2025-06-01#"
#     date_report_match =re.match(r"^[#!]\s*日报#(\d{4}-\d{1,2}-\d{1,2})#", user_input.strip(), re.IGNORECASE)
#     if date_report_match:
#         try:
#             date_str = date_report_match.group(1)
#             date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
#             info = f"用户请求生成{date}日报。"
#         except ValueError:
#             info = date_error
    
#     sale_match =re.match(r"^[#!]\s*销量#(\d{4}-\d{1,2}-\d{1,2})#(\d{4}-\d{1,2}-\d{1,2})#([A-Z0-9]{10})#", user_input.strip(), re.IGNORECASE)
#     if sale_match:
#         try:
#             date_str1 = sale_match.group(1)
#             date_str2 = sale_match.group(2)
#             asin_str = sale_match.group(3)
#             date1 = datetime.datetime.strptime(date_str1, "%Y-%m-%d").date()
#             date2 = datetime.datetime.strptime(date_str2, "%Y-%m-%d").date()
#             info=f"用户请求生成{date1}到{date2} {asin_str}销量。"
#         except ValueError:
#             info = date_error
    
#     date_sale_match =re.match(r"^[#!]\s*销量#(\d{4}-\d{1,2}-\d{1,2})#([A-Z0-9]{10})#", user_input.strip(), re.IGNORECASE)
#     if date_sale_match:
#         try:
#             date_str = date_sale_match.group(1)
#             asin_str = date_sale_match.group(2)
#             date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
#             info=f"用户请求生成{date} {asin_str}销量。"
#         except ValueError:
#             info = date_error
    
#     ranking_match =re.match(r"^[#!]\s*排名#(\d{4}-\d{1,2}-\d{1,2})#(\d{4}-\d{1,2}-\d{1,2})#", user_input.strip(), re.IGNORECASE)
#     if ranking_match:
#         try:
#             date_str1 = ranking_match.group(1)
#             date_str2 = ranking_match.group(2)
#             date1 = datetime.datetime.strptime(date_str1, "%Y-%m-%d").date()
#             date2 = datetime.datetime.strptime(date_str2, "%Y-%m-%d").date()
#             info = f"用户请求生成{date1}到{date2}排名。"
#         except ValueError:
#             info = date_error
    
#     date_ranking_match =re.match(r"^[#!]\s*排名#(\d{4}-\d{1,2}-\d{1,2})#", user_input.strip(), re.IGNORECASE)
#     if date_ranking_match:
#         try:
#             date_str = date_ranking_match.group(1)
#             date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
#             info = f"用户请求生成{date}排名。"
#         except ValueError:
#             info = date_error
    
#     if info=="":
#         info="听不懂"


#     content = json.dumps(
#         {
#             "text": "收到你的消息: "
#             + user_input
#             +"\n"
#             + info
#             # + str(dat)
#         }
#     )
    
#     if data.event.message.chat_type == "p2p":
#         request = (
#             CreateMessageRequest.builder()
#             .receive_id_type("chat_id")
#             .request_body(
#                 CreateMessageRequestBody.builder()
#                 .receive_id(data.event.message.chat_id)
#                 .msg_type("text")
#                 .content(content)
#                 .build()
#             )
#             .build()
#         )

#         # 使用OpenAPI发送消息
#         # Use send OpenAPI to send messages
#         # https://open.feishu.cn/document/uAjLw4CM/ukTMukTMukTM/reference/im-v1/message/create
#         response = client.im.v1.chat.create(request)

#         if not response.success():
#             raise Exception(
#                 f"client.im.v1.chat.create failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}"
#             )
#     else:
#         request: ReplyMessageRequest = (
#             ReplyMessageRequest.builder()
#             .message_id(data.event.message.message_id)
#             .request_body(
#                 ReplyMessageRequestBody.builder()
#                 .content(content)
#                 .msg_type("text")
#                 .build()
#             )
#             .build()
#         )
#         # 使用OpenAPI回复消息
#         # Reply to messages using send OpenAPI
#         # https://open.feishu.cn/document/uAjLw4CM/ukTMukTMukTM/reference/im-v1/message/reply
#         response: ReplyMessageResponse = client.im.v1.message.reply(request)
#         if not response.success():
#             raise Exception(
#                 f"client.im.v1.message.reply failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}"
#             )
#     print(f'[ do_p2_im_message_receive_v1 access ], data: {lark.JSON.marshal(data, indent=4)}')

# def do_message_event(data: lark.CustomizedEvent) -> None:
#     print(f'[ do_customized_event access ], type: message, data: {lark.JSON.marshal(data, indent=4)}')

# event_handler = lark.EventDispatcherHandler.builder("", "") \
#     .register_p2_im_message_receive_v1(do_p2_im_message_receive_v1) \
#     .register_p1_customized_event("这里填入你要自定义订阅的 event 的 key，例如 out_approval", do_message_event) \
#     .build()
# client = lark.Client.builder().app_id(APP_ID).app_secret(APP_SECRET).build()

# def main():
#     cli = lark.ws.Client(APP_ID, APP_SECRET,
#                          event_handler=event_handler,
#                          log_level=lark.LogLevel.DEBUG)
#     cli.start()

# if __name__ == "__main__":
#     main()